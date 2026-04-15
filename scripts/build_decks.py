"""Build Anki-importable CSV decks from Heisig data sources.

Outputs:
  RTH_deck.csv  — Traditional Hanzi only
  RSH_deck.csv  — Simplified Hanzi only
  RTK_deck.csv  — Kanji only
  Ultimate_deck.csv — All 3 merged, one card per unique character
"""
import csv
import json
import re
import openpyxl
from collections import defaultdict

# ── Paths ──────────────────────────────────────────────────────────────
EXCEL = "data/Heisig's Remembering the Kanji vs. Hanzi v27.xlsx"
RSH_JSON = "data/rsh_parsed.json"
IDS_TXT = "data/IDS.TXT"
UNIFIED_MAP = "data/unified_mapping.json"
HUMAN_REVIEW = "data/unmapped_human_reviewed.csv"

# ── IDS operator labels ────────────────────────────────────────────────
IDS_LABELS = {
    "⿰": "left-right",
    "⿱": "top-bottom",
    "⿲": "left-mid-right",
    "⿳": "top-mid-bottom",
    "⿴": "surround",
    "⿵": "surround-open-bottom",
    "⿶": "surround-open-top",
    "⿷": "surround-open-right",
    "⿸": "top-left-wrap",
    "⿹": "top-right-wrap",
    "⿺": "bottom-left-wrap",
    "⿻": "overlaid",
}
IDS_OPERATORS = set("⿰⿱⿲⿳⿴⿵⿶⿷⿸⿹⿺⿻⿼⿽⿾⿿〾")

# ══════════════════════════════════════════════════════════════════════
# 1. Load data sources
# ══════════════════════════════════════════════════════════════════════

# ── RSH parsed JSON ────────────────────────────────────────────────────
with open(RSH_JSON, "r", encoding="utf-8") as f:
    rsh = json.load(f)

heisig_by_char = {}
heisig_by_keyword = {}
# Two passes: keywords first, then aliases (aliases override keywords)
for e in rsh["characters"] + rsh["primitives"]:
    heisig_by_char[e["character"]] = e
    heisig_by_keyword[e["keyword"]] = e["character"]
for e in rsh["characters"] + rsh["primitives"]:
    for a in e["primitive_aliases"]:
        heisig_by_keyword[a] = e["character"]

# ── Unified mapping (char -> name) ────────────────────────────────────
with open(UNIFIED_MAP, "r", encoding="utf-8") as f:
    unified = json.load(f)

# ── Human review CSV (extra name overrides) ────────────────────────────
human_names = {}
with open(HUMAN_REVIEW, "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    for row in reader:
        comp = row["component"]
        name = row.get("your_heisig_name", "").strip()
        if name:
            human_names[comp] = name

# ── Radical mappings (same as recursive_decompose_v2) ──────────────────
# Maps variant forms to their "parent" for keyword lookup
radical_map = {
    '訁': '言', '糹': '糸', '釒': '金', '𥫗': '竹', '刂': '刀',
    '彳': '行', '𤣩': '玉', '𧾷': '足', '罒': '网', '乚': '乙',
    '飠': '食', '爫': '爪', '虍': '虎', '𧘇': '衣', '龶': '生',
    '𦍌': '羊', '亍': '行', '牜': '牛', '覀': '西', '丬': '爿',
    '䒑': '丷', '亻': '人', '氵': '水', '扌': '手', '忄': '心',
    '犭': '犬', '礻': '示', '衤': '衣', '灬': '火', '⺌': '小',
    '⺊': '卜', '讠': '言', '钅': '金', '饣': '食', '纟': '糸',
    '贝': '貝', '车': '車', '见': '見', '门': '門', '鱼': '魚',
    '马': '馬', '鸟': '鳥', '页': '頁', '风': '風', '⺝': '月',
    '⺼': '月', '⺶': '羊', '⺀': '八', '⺄': '乙', '⺆': '冂',
    '⺈': '刀',
}

# Traditional → Simplified mappings (for when simplified is the Heisig entry)
# This handles cases like 門→门 where Heisig uses simplified
trad_to_simp_map = {
    '門': '门', '貝': '贝', '車': '车', '見': '见', '魚': '鱼',
    '馬': '马', '鳥': '鸟', '頁': '页', '風': '风', '言': '讠',
    '金': '钅', '食': '饣', '糸': '纟',
}

for variant, parent in radical_map.items():
    if variant not in heisig_by_char and parent in heisig_by_char:
        heisig_by_char[variant] = {
            "character": variant,
            "keyword": heisig_by_char[parent]["keyword"],
            "type": "radical_variant",
            "primitive_aliases": heisig_by_char[parent]["primitive_aliases"],
            "components": [],
            "variant_of": parent,
        }

# Also add reverse mappings: traditional → simplified when simplified has the keyword
for trad, simp in trad_to_simp_map.items():
    if trad not in heisig_by_char and simp in heisig_by_char:
        heisig_by_char[trad] = {
            "character": trad,
            "keyword": heisig_by_char[simp]["keyword"],
            "type": "trad_variant",
            "primitive_aliases": heisig_by_char[simp]["primitive_aliases"],
            "components": [],
            "variant_of": simp,
        }

# ── Excel workbook ─────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL)
ws = wb["RTH+RSH+RTK"]

# Trad/Kanji -> Simplified variant mapping into heisig_by_char
for row in ws.iter_rows(min_row=2, values_only=True):
    th, sh, k = row[3], row[4], row[5]
    for char in [th, k]:
        if char and sh and char != sh and char not in heisig_by_char and sh in heisig_by_char:
            heisig_by_char[char] = {
                "character": char,
                "keyword": heisig_by_char[sh]["keyword"],
                "type": "variant",
                "primitive_aliases": heisig_by_char[sh]["primitive_aliases"],
                "components": [],
                "variant_of": sh,
            }

# ── IDS data ───────────────────────────────────────────────────────────
# Numbered component definitions
numbered_components = {}
with open(IDS_TXT, "r", encoding="utf-8-sig") as f:
    for line in f:
        m = re.match(r'^#\s+\{(\d+)\}\s+(.+)', line)
        if m:
            num = int(m.group(1))
            desc = m.group(2).strip()
            parts = desc.split('\t')
            description = parts[0].strip()
            expansion = parts[-1].strip() if len(parts) > 1 else None
            if expansion == '？':
                expansion = None
            numbered_components[num] = {"description": description, "expansion": expansion}

# IDS sequences
ids_map = {}
with open(IDS_TXT, "r", encoding="utf-8-sig") as f:
    for line in f:
        if line.startswith("#") or line.strip() == "":
            continue
        parts = line.strip().split("\t")
        if len(parts) >= 3:
            ids_map[parts[1]] = parts[2:]

# ── CC-CEDICT from Excel ──────────────────────────────────────────────
ws_cedict = wb["CC-CEDICT"]
# Group readings by (TH, SH) pair
cedict_by_char = defaultdict(list)  # char -> [(pinyin, definition)]
for row in ws_cedict.iter_rows(min_row=2, values_only=True):
    th, sh, pinyin, defn = row[0], row[1], row[2], row[3]
    if not pinyin:
        continue
    for char in [th, sh]:
        if char:
            cedict_by_char[char].append((pinyin, defn or ""))

print(f"Loaded: {len(heisig_by_char)} Heisig chars, {len(ids_map)} IDS entries, "
      f"{len(cedict_by_char)} CC-CEDICT chars, {len(unified)} unified mappings")

# ══════════════════════════════════════════════════════════════════════
# 2. Helper functions
# ══════════════════════════════════════════════════════════════════════

def get_heisig_name(char):
    """Get the Heisig primitive/keyword name for a character."""
    if char in human_names:
        return human_names[char]
    if char in heisig_by_char:
        e = heisig_by_char[char]
        if e["primitive_aliases"]:
            return e["primitive_aliases"][0]
        return e["keyword"]
    if char in unified:
        return unified[char]["name"]
    return None


def tokenize_ids(ids_str):
    cleaned = re.sub(r'\$\([^)]*\)', '', ids_str)
    cleaned = cleaned.replace('^', '').strip()
    tokens = []
    i = 0
    while i < len(cleaned):
        ch = cleaned[i]
        if ch == '{':
            j = cleaned.index('}', i)
            tokens.append(('numbered', int(cleaned[i+1:j])))
            i = j + 1
        elif ch in ' \t\r\n()':
            i += 1
        else:
            tokens.append(('char', ch))
            i += 1
    return tokens


def parse_ids(ids_str):
    tokens = tokenize_ids(ids_str)
    pos = [0]
    def parse_next():
        if pos[0] >= len(tokens):
            return None
        tok_type, tok_val = tokens[pos[0]]
        pos[0] += 1
        if tok_type == 'char' and tok_val in IDS_OPERATORS:
            n = 3 if tok_val in '⿲⿳' else (1 if tok_val == '〾' else 2)
            children = []
            for _ in range(n):
                child = parse_next()
                if child is not None:
                    children.append(child)
            return ('op', tok_val, children)
        elif tok_type == 'numbered':
            return ('numbered', tok_val)
        elif tok_type == 'char':
            return ('char', tok_val)
        return parse_next()
    return parse_next()


def recursive_decompose(char, depth=0, max_depth=10, seen=None, is_root=True):
    """Decompose a character into named Heisig components."""
    if seen is None:
        seen = set()
    if char in seen or depth > max_depth:
        return {"char": char, "name": get_heisig_name(char)}
    seen = seen | {char}
    name = get_heisig_name(char)

    # Heisig decomposition from XML (direct only — not via variant,
    # since trad/kanji variants often have different internal structure)
    if char in heisig_by_char and heisig_by_char[char].get("components"):
        e = heisig_by_char[char]
        children = []
        char_to_idx = {}  # comp_char -> index in children, for dedup
        for comp_name in e["components"]:
            comp_char = heisig_by_keyword.get(comp_name)
            if comp_char and comp_char in char_to_idx:
                # Same character cited twice (e.g. "early" and "sunflower" both = 早)
                # Merge alt name onto existing child
                idx = char_to_idx[comp_char]
                children[idx].setdefault("alt_names", []).append(comp_name)
            else:
                child = {"char": comp_char or "?", "name": comp_name}
                if comp_char:
                    char_to_idx[comp_char] = len(children)
                children.append(child)
        return {"char": char, "name": name, "source": "heisig", "children": children}

    # If this is a sub-component (not root) and it has a name
    # (from Heisig, human review, or unified mapping), treat as atomic
    if not is_root and name:
        return {"char": char, "name": name, "source": "named_atomic"}

    # For the root character or unnamed characters, try IDS decomposition
    if char in ids_map:
        tree = parse_ids(ids_map[char][0])
        return _decompose_ids_tree(tree, depth, max_depth, seen)

    # If no IDS available, try the variant's decomposition as fallback
    if char in heisig_by_char:
        variant_of = heisig_by_char[char].get("variant_of")
        if variant_of and variant_of in heisig_by_char and heisig_by_char[variant_of].get("components"):
            e = heisig_by_char[variant_of]
            children = []
            char_to_idx = {}
            for comp_name in e["components"]:
                comp_char = heisig_by_keyword.get(comp_name)
                if comp_char and comp_char in char_to_idx:
                    children[char_to_idx[comp_char]].setdefault("alt_names", []).append(comp_name)
                else:
                    child = {"char": comp_char or "?", "name": comp_name}
                    if comp_char:
                        char_to_idx[comp_char] = len(children)
                    children.append(child)
            return {"char": char, "name": name, "source": "heisig_variant", "children": children}

    # Heisig atomic or mapped — stop
    if name:
        return {"char": char, "name": name, "source": "heisig_atomic"}

    return {"char": char, "name": None, "source": "unknown"}


def _decompose_ids_tree(tree, depth, max_depth, seen):
    if tree is None:
        return {"char": "?", "name": None, "source": "parse_error"}
    kind = tree[0]
    if kind == 'char':
        return recursive_decompose(tree[1], depth + 1, max_depth, seen, is_root=False)
    elif kind == 'numbered':
        num = tree[1]
        comp = numbered_components.get(num, {})
        expansion = comp.get("expansion")
        if expansion:
            subtree = parse_ids(expansion)
            return _decompose_ids_tree(subtree, depth + 1, max_depth, seen)
        return {"char": f"{{{num}}}", "name": None, "source": "numbered_component"}
    elif kind == 'op':
        op = tree[1]
        children = [_decompose_ids_tree(c, depth + 1, max_depth, seen) for c in tree[2]]
        return {"char": "".join(c.get("char", "?") for c in children),
                "name": None, "source": "ids", "operator": op, "children": children}
    return {"char": "?", "name": None, "source": "parse_error"}


def collect_leaves(node, is_root=True):
    """Get named component names from a decomposition tree.
    Stops at any named node (doesn't recurse into its sub-components)."""
    # Named non-root node: stop here
    if not is_root and node.get("name"):
        alt_names = node.get("alt_names", [])
        if alt_names:
            return [node["name"] + " / " + " / ".join(alt_names)]
        return [node["name"]]
    if "children" not in node:
        name = node.get("name")
        return [name] if name else [node.get("char", "?")]
    result = []
    for child in node["children"]:
        result.extend(collect_leaves(child, is_root=False))
    return result


def collect_leaf_details(node, is_root=True):
    """Get (char, name) pairs for named components.
    Stops at any named node (doesn't recurse into its sub-components)."""
    if not is_root and node.get("name"):
        return [(node.get("char", "?"), node["name"])]
    if "children" not in node:
        return [(node.get("char", "?"), node.get("name"))]
    result = []
    for child in node["children"]:
        result.extend(collect_leaf_details(child, is_root=False))
    return result


def get_raw_ids(char):
    """Get the cleaned raw IDS string for a character."""
    if char in ids_map:
        raw = ids_map[char][0]
        return re.sub(r'\$\([^)]*\)', '', raw).replace('^', '').strip()
    return ""


def get_top_operator(char):
    """Get the top-level IDS operator for a character's spatial layout."""
    cleaned = get_raw_ids(char)
    if cleaned and cleaned[0] in IDS_LABELS:
        op = cleaned[0]
        return f"{op} ({IDS_LABELS[op]})"
    return ""


def format_component_html(ch, keyword, aliases, decomp_name):
    """Format a component as styled HTML: blue char, gray keyword."""
    # Build alias suffix if needed
    alias_suffix = ""
    if decomp_name.lower() != keyword.lower() and decomp_name.lower() in [a.lower() for a in aliases]:
        alias_suffix = f' <span style="color:#999">(alias: {decomp_name})</span>'
    elif aliases and any(a.lower() != keyword.lower() for a in aliases):
        other_aliases = [a for a in aliases if a.lower() != keyword.lower()]
        if other_aliases:
            alias_suffix = f' <span style="color:#999">(alias: {", ".join(other_aliases)})</span>'

    return (f'<span style="color:#1a5276">{ch}</span> '
            f'<span style="color:#666">{keyword}</span>{alias_suffix}')


def format_reading(char):
    """Return CC-CEDICT pinyin readings for a character, no definitions.

    Multiple readings separated by ' / '. Surname-only readings are filtered
    out when other readings exist. Max 4 readings returned.
    """
    entries = cedict_by_char.get(char, [])
    if not entries:
        return ""

    by_pinyin = defaultdict(list)
    for pinyin, defn in entries:
        by_pinyin[pinyin].append(defn or "")

    filtered = []
    for pinyin, defns in by_pinyin.items():
        # Skip surname-only readings when other readings exist
        if len(by_pinyin) > 1:
            all_meanings = []
            for d in defns:
                for m in d.split("/"):
                    m = m.strip()
                    if m:
                        all_meanings.append(m)
            if all_meanings and all(
                "surname" in m.lower() or "name" in m.lower()
                for m in all_meanings
            ):
                continue
        filtered.append(pinyin)
        if len(filtered) >= 4:
            break

    return " / ".join(filtered)


# ══════════════════════════════════════════════════════════════════════
# 3. Build card data from Excel
# ══════════════════════════════════════════════════════════════════════

# Collect all Excel row data
excel_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    excel_rows.append(row)

# Known Excel data errors: patch before any processing
# 愣 (dumbfounded) has no traditional form; TH=傻 in its row is incorrect
excel_rows = [
    tuple(None if (i == 3 and row[3] == '傻' and row[4] == '愣') else v
          for i, v in enumerate(row))
    for row in excel_rows
]

# Build card dicts keyed by character
cards = {}  # char -> card dict

def ensure_card(char):
    if char not in cards:
        cards[char] = {
            "character": char,
            "keyword": "",
            "primitive_meanings": "",
            "RTH_number": "",
            "RSH_number": "",
            "RTK_number": "",
            "reading": "",
            "decomposition": "",
            "spatial": "",
            "components_detail": "",
            "tags": [],
            "books": set(),  # internal tracking
        }
    return cards[char]


for row in excel_rows:
    rth_num, rsh_num, rtk_num = row[0], row[1], row[2]
    th, sh, k = row[3], row[4], row[5]
    rth_kw, rsh_kw, rtk_kw = row[7], row[8], row[9]
    rth_read, th_read = row[10], row[11]
    rth_lesson, rsh_lesson, rtk_lesson = row[12], row[13], row[14]

    # Parse RSH number (can be "ch # 0041" format or numeric)
    def parse_num(val):
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return int(val)
        m = re.match(r'ch\s*#\s*(\d+)', str(val))
        if m:
            return int(m.group(1))
        return None

    rth_n = parse_num(rth_num)
    rsh_n = parse_num(rsh_num)
    rtk_n = parse_num(rtk_num)

    # Process each book's character
    book_entries = [
        (th, rth_n, rth_kw, rth_lesson, "RTH"),
        (sh, rsh_n, rsh_kw, rsh_lesson, "RSH"),
        (k, rtk_n, rtk_kw, rtk_lesson, "RTK"),
    ]

    for char, num, kw, lesson, book in book_entries:
        if not char:
            continue
        card = ensure_card(char)

        if num:
            card["books"].add(book)
            card[f"{book}_number"] = str(num)

        if kw:
            card[f"{book}_keyword"] = kw
        if kw and not card["keyword"]:
            card["keyword"] = kw

        if lesson:
            # Convert "RSH1-L01" -> "RSH1::L01" for Anki nested tags
            card["tags"].append(lesson.replace("-", "::"))


# ── Build variant mappings ────────────────────────────────────────────
# Map each character to its variants in other systems
char_variants = {}  # char -> {"th": ..., "sh": ..., "k": ...}
for row in excel_rows:
    th, sh, k = row[3], row[4], row[5]
    for char in [th, sh, k]:
        if char and char not in char_variants:
            char_variants[char] = {"th": th, "sh": sh, "k": k}


def format_variants(char, for_book):
    """Format variant string for a character, relative to a book.

    for_book: 'RTH', 'RSH', or 'RTK' - determines which variants to show.
    """
    if char not in char_variants:
        return ""
    v = char_variants[char]
    th, sh, k = v["th"], v["sh"], v["k"]

    parts = []
    if for_book == "RTH":  # Traditional - show Simplified and Japanese
        if sh and sh != char:
            parts.append(f"SH: {sh}")
        if k and k != char:
            parts.append(f"JP: {k}")
    elif for_book == "RSH":  # Simplified - show Traditional and Japanese
        if th and th != char:
            parts.append(f"TH: {th}")
        if k and k != char:
            parts.append(f"JP: {k}")
    elif for_book == "RTK":  # Japanese - show Traditional and Simplified
        if th and th != char:
            parts.append(f"TH: {th}")
        if sh and sh != char:
            parts.append(f"SH: {sh}")

    if not parts:
        return ""
    return " | ".join(parts)


# ══════════════════════════════════════════════════════════════════════
# 4. Enrich cards: decomposition, spatial, readings, primitives
# ══════════════════════════════════════════════════════════════════════

for char, card in cards.items():
    # Deck column: which book(s) this character belongs to
    card["deck"] = " ".join(sorted(card["books"])) if card["books"] else ""

    # Reading from CC-CEDICT (skip for RTK-only / kanji-only)
    if "RTH" in card["books"] or "RSH" in card["books"]:
        card["reading"] = format_reading(char)

    # Decomposition + components_detail
    tree = recursive_decompose(char)
    leaves = collect_leaves(tree)
    leaf_details = collect_leaf_details(tree)

    if tree.get("source") == "heisig" or (tree.get("children") and len(leaves) > 0):
        card["decomposition"] = " + ".join(leaves)
        detail_parts = []
        seen_detail = set()
        for ch, decomp_name in leaf_details:
            if decomp_name and ch != "?" and ch not in seen_detail:
                seen_detail.add(ch)
                if ch in heisig_by_char:
                    keyword = heisig_by_char[ch].get("keyword", decomp_name)
                    aliases = heisig_by_char[ch].get("primitive_aliases", [])
                else:
                    keyword = decomp_name
                    aliases = []
                detail_parts.append(format_component_html(ch, keyword, aliases, decomp_name))
        card["components_detail"] = "<br>".join(detail_parts)
    elif tree.get("source") == "heisig_atomic":
        card["decomposition"] = ""  # atomic, no sub-components
        card["components_detail"] = ""

    # Spatial from IDS
    card["spatial"] = get_top_operator(char)

    # Raw IDS string
    card["ids"] = get_raw_ids(char)

    # Primitive aliases (shown on card as dimmed line below keyword)
    if char in heisig_by_char:
        entry = heisig_by_char[char]
        aliases = entry.get("primitive_aliases", [])
        kw_lower = card["keyword"].lower()
        distinct = [a for a in aliases if a.lower() != kw_lower]
        card["primitive_meanings"] = " · ".join(distinct) if distinct else ""
    else:
        card["primitive_meanings"] = ""

    # Tags: add primitive tag if applicable
    if char in heisig_by_char:
        entry = heisig_by_char[char]
        if entry.get("type") == "primitive" or entry.get("primitive_aliases"):
            card["tags"].append("primitive")

    card["tags"] = " ".join(sorted(set(card["tags"])))


# Patch primitive cards that came from Excel with no lesson tag:
# use lesson from rsh["primitives"] if available.
prim_lesson_by_char = {
    e["character"]: f"RSH{e['book']}::L{e['lesson']:02d}"
    for e in rsh["primitives"]
    if e.get("book") and e.get("lesson")
}
for char, card in cards.items():
    tags = card.get("tags", "")
    if "primitive" not in tags:
        continue
    has_lesson = any(t.startswith("RSH") for t in tags.split())
    if not has_lesson and char in prim_lesson_by_char:
        card["tags"] = " ".join(sorted(set(tags.split() + [prim_lesson_by_char[char]])))


def parse_lesson_tag(tag):
    """Parse 'RTH1::L03' -> (1, 3) for sorting."""
    m = re.match(r'[A-Z]+(\d+)::L(\d+)', tag)
    return (int(m.group(1)), int(m.group(2))) if m else (99, 99)


# Build per-primitive first-appearance lesson for RTH and RTK:
# scan all Heisig-decomposed characters and find the earliest lesson
# in which each primitive appears as a component.
prim_first_lesson = {}  # prim_char -> {"RTH": tag, "RTK": tag}

for char, card in cards.items():
    if char not in heisig_by_char:
        continue
    components = heisig_by_char[char].get("components", [])
    if not components:
        continue
    tags = card.get("tags", "").split()
    book_tags = {
        "RTH": next((t for t in tags if t.startswith("RTH") and "::" in t), None),
        "RTK": next((t for t in tags if t.startswith("RTK") and "::" in t), None),
    }
    for comp_name in components:
        prim_char = heisig_by_keyword.get(comp_name)
        if not prim_char or prim_char not in heisig_by_char:
            continue
        prim_entry = heisig_by_char[prim_char]
        if prim_entry.get("type") != "primitive" and not prim_entry.get("primitive_aliases"):
            continue
        if prim_char not in prim_first_lesson:
            prim_first_lesson[prim_char] = {}
        for book, tag in book_tags.items():
            if tag:
                existing = prim_first_lesson[prim_char].get(book)
                if not existing or parse_lesson_tag(tag) < parse_lesson_tag(existing):
                    prim_first_lesson[prim_char][book] = tag


# ══════════════════════════════════════════════════════════════════════
# 5. Add standalone primitive cards
# ══════════════════════════════════════════════════════════════════════

primitives_added = 0
for entry in rsh["primitives"]:
    char = entry["character"]
    if char in cards:
        # Card exists (e.g. from Excel as RTK-only). Patch in RSH primitive identity
        # so it appears in the RSH deck with the correct keyword and lesson tag.
        existing = cards[char]
        lesson_tag = ""
        if entry.get("book") and entry.get("lesson"):
            lesson_tag = f"RSH{entry['book']}::L{entry['lesson']:02d}"
        books_to_add = ["RSH"] + [b for b in ("RTH", "RTK") if b in prim_first_lesson.get(char, {})]
        for book in books_to_add:
            if book not in existing.get("deck", ""):
                existing["deck"] = (existing.get("deck", "") + " " + book).strip()
        if not existing.get("RSH_keyword"):
            existing["RSH_keyword"] = entry["keyword"]
        tags = set(existing.get("tags", "").split())
        tags.add("primitive")
        if lesson_tag:
            tags.add(lesson_tag)
        for book, book_tag in prim_first_lesson.get(char, {}).items():
            tags.add(book_tag)
        existing["tags"] = " ".join(sorted(tags))
        continue
    # Standalone primitive not in any book's character list
    tree = recursive_decompose(char)
    leaves = collect_leaves(tree)
    leaf_details = collect_leaf_details(tree)

    aliases = entry["primitive_aliases"]
    kw = entry["keyword"]

    lesson_tag = ""
    if entry.get("book") and entry.get("lesson"):
        lesson_tag = f"RSH{entry['book']}::L{entry['lesson']:02d}"

    extra_tags = [t for t in prim_first_lesson.get(char, {}).values()]
    tags = " ".join(sorted(filter(None, [lesson_tag] + extra_tags + ["primitive"])))

    distinct_aliases = [a for a in aliases if a.lower() != kw.lower()]
    card = {
        "character": char,
        "keyword": kw,
        "primitive_meanings": " · ".join(distinct_aliases),
        "RTH_number": "",
        "RSH_number": "",
        "RTK_number": "",
        "reading": format_reading(char),
        "decomposition": " + ".join(leaves) if tree.get("children") else "",
        "spatial": get_top_operator(char),
        "ids": get_raw_ids(char),
        "components_detail": "",
        "deck": " ".join(["RSH"] + [b for b in ("RTH", "RTK") if b in prim_first_lesson.get(char, {})]),
        "tags": tags,
    }
    if tree.get("children"):
        detail_parts = []
        seen_detail = set()
        for ch, decomp_name in leaf_details:
            if decomp_name and ch != "?" and ch not in seen_detail:
                seen_detail.add(ch)
                if ch in heisig_by_char:
                    keyword = heisig_by_char[ch].get("keyword", decomp_name)
                    aliases = heisig_by_char[ch].get("primitive_aliases", [])
                else:
                    keyword = decomp_name
                    aliases = []
                detail_parts.append(format_component_html(ch, keyword, aliases, decomp_name))
        card["components_detail"] = "<br>".join(detail_parts)

    cards[char] = card
    primitives_added += 1

# Also add character-primitives that might only be in RSH XML but not Excel
for entry in rsh["characters"]:
    char = entry["character"]
    if char in cards:
        continue
    if not entry["primitive_aliases"]:
        continue
    tree = recursive_decompose(char)
    leaves = collect_leaves(tree)
    leaf_details = collect_leaf_details(tree)

    lesson_tag = ""
    if entry.get("book") and entry.get("lesson"):
        lesson_tag = f"RSH{entry['book']}::L{entry['lesson']:02d}"

    kw = entry["keyword"]
    aliases = entry["primitive_aliases"]
    distinct_aliases = [a for a in aliases if a.lower() != kw.lower()]
    card = {
        "character": char,
        "keyword": kw,
        "primitive_meanings": " · ".join(distinct_aliases),
        "RTH_number": "",
        "RSH_number": str(entry["number"]) if entry.get("number") else "",
        "RTK_number": "",
        "reading": format_reading(char),
        "decomposition": " + ".join(leaves) if tree.get("children") else "",
        "spatial": get_top_operator(char),
        "ids": get_raw_ids(char),
        "components_detail": "",
        "deck": " ".join(["RSH"] + [b for b in ("RTH", "RTK") if b in prim_first_lesson.get(char, {})]),
        "tags": " ".join(sorted(filter(None, [lesson_tag] + list(prim_first_lesson.get(char, {}).values()) + ["primitive"]))),
    }
    if tree.get("children"):
        detail_parts = []
        seen_detail = set()
        for ch, decomp_name in leaf_details:
            if decomp_name and ch != "?" and ch not in seen_detail:
                seen_detail.add(ch)
                if ch in heisig_by_char:
                    keyword = heisig_by_char[ch].get("keyword", decomp_name)
                    aliases = heisig_by_char[ch].get("primitive_aliases", [])
                else:
                    keyword = decomp_name
                    aliases = []
                detail_parts.append(format_component_html(ch, keyword, aliases, decomp_name))
        card["components_detail"] = "<br>".join(detail_parts)

    cards[char] = card
    primitives_added += 1

print(f"Standalone primitives added: {primitives_added}")

# ══════════════════════════════════════════════════════════════════════
# 6. Output CSVs
# ══════════════════════════════════════════════════════════════════════

COLUMNS = ["character", "keyword", "primitive_meanings", "RTH_number", "RSH_number", "RTK_number",
           "reading", "decomposition", "spatial", "ids", "components_detail", "variants", "deck", "tags"]


def write_deck(filename, filter_fn, tag_prefix=None, for_book=None):
    """Write a CSV deck, filtering cards by filter_fn.

    If tag_prefix is provided, only include tags starting with that prefix
    (plus 'primitive' tag). This ensures single-book decks don't include
    tags from other books.

    If for_book is provided (RTH/RSH/RTK), populate variants field and
    clear irrelevant book numbers.
    """
    rows = []
    for char, card in sorted(cards.items(), key=lambda x: x[0]):
        if filter_fn(card):
            row_data = []
            for col in COLUMNS:
                val = card.get(col, "")
                # Filter tags for single-book decks
                if col == "tags" and tag_prefix and val:
                    tags = val.split()
                    filtered = [t for t in tags if t.startswith(tag_prefix) or t == "primitive"]
                    val = " ".join(filtered)
                # Add variants for single-book decks
                if col == "variants" and for_book:
                    val = format_variants(char, for_book)
                # Use book-specific keyword for single-book decks
                if col == "keyword" and for_book:
                    book_kw = card.get(f"{for_book}_keyword", "")
                    if book_kw:
                        val = book_kw
                # Clear irrelevant book numbers for single-book decks
                if for_book:
                    if col == "RTH_number" and for_book != "RTH":
                        val = ""
                    elif col == "RSH_number" and for_book != "RSH":
                        val = ""
                    elif col == "RTK_number" and for_book != "RTK":
                        val = ""
                row_data.append(val)
            rows.append(row_data)

    with open(filename, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(COLUMNS)
        writer.writerows(rows)

    return len(rows)


# Collect which characters belong to which books from Excel
rth_chars = set()
rsh_chars = set()
rtk_chars = set()
for row in excel_rows:
    if row[0] and row[3]:  # RTH number + TH char
        rth_chars.add(row[3])
    if row[1] and row[4]:  # RSH number + SH char
        rsh_chars.add(row[4])
    if row[2] and row[5]:  # RTK number + K char
        rtk_chars.add(row[5])

is_primitive = lambda c: "primitive" in c.get("tags", "")
# Primitives should only be included if they belong to the target book
is_rth_primitive = lambda c: is_primitive(c) and "RTH" in c.get("deck", "")
is_rsh_primitive = lambda c: is_primitive(c) and "RSH" in c.get("deck", "")
is_rtk_primitive = lambda c: is_primitive(c) and "RTK" in c.get("deck", "")

n_rth = write_deck("RTH_deck.csv", lambda c: c["character"] in rth_chars or is_rth_primitive(c), tag_prefix="RTH", for_book="RTH")
n_rsh = write_deck("RSH_deck.csv", lambda c: c["character"] in rsh_chars or is_rsh_primitive(c), tag_prefix="RSH", for_book="RSH")
n_rtk = write_deck("RTK_deck.csv", lambda c: c["character"] in rtk_chars or is_rtk_primitive(c), tag_prefix="RTK", for_book="RTK")


# ══════════════════════════════════════════════════════════════════════
# 6b. Build family-based Ultimate deck
# ══════════════════════════════════════════════════════════════════════

ULTIMATE_COLUMNS = [
    "character", "keyword", "RTH_number", "RSH_number", "RTK_number",
    "reading", "decomposition", "spatial", "ids", "components_detail",
    "simplified", "simplified_reading", "simplified_decomposition", "simplified_components",
    "japanese", "japanese_decomposition", "japanese_components",
    "deck", "tags"
]


def get_card_decomposition(char):
    """Get decomposition data for a character."""
    if char in cards:
        return {
            "decomposition": cards[char].get("decomposition", ""),
            "components_detail": cards[char].get("components_detail", ""),
            "reading": cards[char].get("reading", ""),
        }
    # Compute on the fly if not in cards dict
    tree = recursive_decompose(char)
    leaves = collect_leaves(tree)
    leaf_details = collect_leaf_details(tree)

    decomp = " + ".join(leaves) if tree.get("children") else ""
    detail_parts = []
    seen = set()
    for ch, decomp_name in leaf_details:
        if decomp_name and ch != "?" and ch not in seen:
            seen.add(ch)
            if ch in heisig_by_char:
                keyword = heisig_by_char[ch].get("keyword", decomp_name)
                aliases = heisig_by_char[ch].get("primitive_aliases", [])
            else:
                keyword = decomp_name
                aliases = []
            detail_parts.append(format_component_html(ch, keyword, aliases, decomp_name))

    return {
        "decomposition": decomp,
        "components_detail": "<br>".join(detail_parts),
        "reading": format_reading(char),
    }


def write_ultimate_deck(filename):
    """Write family-based Ultimate deck with variant fields."""
    families = []
    seen_canonical = {}  # canonical char -> index in families list

    # Build families from Excel rows
    for row in excel_rows:
        rth_num, rsh_num, rtk_num = row[0], row[1], row[2]
        th, sh, k = row[3], row[4], row[5]
        rth_kw, rsh_kw, rtk_kw = row[7], row[8], row[9]

        # Skip empty rows
        if not any([th, sh, k]):
            continue

        # Pick canonical: prefer TH, then SH, then K
        canonical = th or sh or k
        canonical_card = cards.get(canonical, {})

        # Get canonical data
        family = {
            "character": canonical,
            "keyword": canonical_card.get("keyword", rth_kw or rsh_kw or rtk_kw or ""),
            "RTH_number": canonical_card.get("RTH_number", ""),
            "RSH_number": canonical_card.get("RSH_number", ""),
            "RTK_number": canonical_card.get("RTK_number", ""),
            "reading": canonical_card.get("reading", ""),
            "decomposition": canonical_card.get("decomposition", ""),
            "spatial": canonical_card.get("spatial", ""),
            "ids": canonical_card.get("ids", ""),
            "components_detail": canonical_card.get("components_detail", ""),
            "simplified": "",
            "simplified_reading": "",
            "simplified_decomposition": "",
            "simplified_components": "",
            "japanese": "",
            "japanese_decomposition": "",
            "japanese_components": "",
            "deck": canonical_card.get("deck", ""),
            "tags": canonical_card.get("tags", ""),
        }

        # Add simplified variant if different from canonical
        if sh and sh != canonical:
            simp_data = get_card_decomposition(sh)
            family["simplified"] = sh
            family["simplified_reading"] = simp_data["reading"]
            family["simplified_decomposition"] = simp_data["decomposition"]
            family["simplified_components"] = simp_data["components_detail"]

        # Add Japanese variant if different from canonical and different from simplified
        if k and k != canonical and k != family["simplified"]:
            jp_data = get_card_decomposition(k)
            family["japanese"] = k
            family["japanese_decomposition"] = jp_data["decomposition"]
            family["japanese_components"] = jp_data["components_detail"]

        # Merge into existing family if this canonical was already seen
        if canonical in seen_canonical:
            existing = families[seen_canonical[canonical]]
            # Merge simplified: take whichever row has it
            if not existing["simplified"] and family["simplified"]:
                existing["simplified"] = family["simplified"]
                existing["simplified_reading"] = family["simplified_reading"]
                existing["simplified_decomposition"] = family["simplified_decomposition"]
                existing["simplified_components"] = family["simplified_components"]
            # Merge japanese: prefer a value that differs from simplified
            new_jp = family["japanese"]
            existing_jp = existing["japanese"]
            existing_simp = existing["simplified"]
            if new_jp and new_jp != existing_simp:
                if not existing_jp or existing_jp == existing_simp:
                    existing["japanese"] = new_jp
                    existing["japanese_decomposition"] = family["japanese_decomposition"]
                    existing["japanese_components"] = family["japanese_components"]
        else:
            seen_canonical[canonical] = len(families)
            families.append(family)

    # Also add standalone primitives
    for char, card in cards.items():
        if "primitive" not in card.get("tags", ""):
            continue
        # Check if already covered by a family
        already_covered = any(
            f["character"] == char or f["simplified"] == char or f["japanese"] == char
            for f in families
        )
        if already_covered:
            continue

        families.append({
            "character": char,
            "keyword": card.get("keyword", ""),
            "RTH_number": card.get("RTH_number", ""),
            "RSH_number": card.get("RSH_number", ""),
            "RTK_number": card.get("RTK_number", ""),
            "reading": card.get("reading", ""),
            "decomposition": card.get("decomposition", ""),
            "spatial": card.get("spatial", ""),
            "ids": card.get("ids", ""),
            "components_detail": card.get("components_detail", ""),
            "simplified": "",
            "simplified_reading": "",
            "simplified_decomposition": "",
            "simplified_components": "",
            "japanese": "",
            "japanese_decomposition": "",
            "japanese_components": "",
            "deck": card.get("deck", ""),
            "tags": card.get("tags", ""),
        })

    # Write CSV
    with open(filename, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(ULTIMATE_COLUMNS)
        for fam in families:
            if not fam.get("keyword", "").strip():
                continue
            writer.writerow([fam.get(col, "") for col in ULTIMATE_COLUMNS])

    # Count stats
    n_total = len([f for f in families if f.get("keyword", "").strip()])
    n_simp = len([f for f in families if f.get("simplified")])
    n_jp = len([f for f in families if f.get("japanese")])

    return n_total, n_simp, n_jp


n_ult, n_simp_cards, n_jp_cards = write_ultimate_deck("Ultimate_deck.csv")

# ══════════════════════════════════════════════════════════════════════
# 7. Summary
# ══════════════════════════════════════════════════════════════════════

print(f"\n{'='*60}")
print(f"  RTH_deck.csv:      {n_rth} cards")
print(f"  RSH_deck.csv:      {n_rsh} cards")
print(f"  RTK_deck.csv:      {n_rtk} cards")
print(f"  Ultimate_deck.csv: {n_ult} families")
print(f"    → Simplified variant cards: {n_simp_cards}")
print(f"    → Japanese variant cards:   {n_jp_cards}")
print(f"{'='*60}")

# Spot checks
for ch, expected in [("虎", "magic wand"), ("國", "pent in")]:
    card = cards.get(ch)
    if card:
        decomp = card["decomposition"]
        has = expected in decomp if decomp else False
        status = "OK" if has else "MISSING"
        print(f"  Spot check {ch}: decomposition = '{decomp}' [{status}]")

# Primitive count
prim_count = sum(1 for c in cards.values() if "primitive" in c.get("tags", ""))
print(f"  Primitive-tagged cards: {prim_count}")
print(f"  Total unique characters: {len(cards)}")
